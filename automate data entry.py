import pandas as pd
import numpy as np
from datetime import datetime  , timedelta
import datetime as dt
import math
import openpyxl
from openpyxl.styles import PatternFill,Border,Side
from openpyxl.utils import get_column_letter

"""--------------------------IMP-----------------------IMP----------------------------IMP------
make an empty file  And save it with the with name nd change the name in the code  below both for line 13 
and 74 ,135,194

--------------------------IMP----------------------IMP---------------------------IMP----------"""

#enter  directory of file 
df=pd.read_excel(r"C:\Users\Tanishqa\Desktop\data\file1.xlsx")
df['Date'] = pd.to_datetime(df['Date'])
today_date = datetime.now().strftime('%dth_%B')


top_7_dates = df['Date'].nlargest(7)

day_names = top_7_dates.dt.strftime('%A')

def agent(df,sheet):
        same_date_values = df.loc[df['Date'].isin(top_7_dates), 'AGENT_BOOKING'].tolist()
        seven_days_ago = top_7_dates - timedelta(days=7)
        seven_days_values = df.loc[df['Date'].isin(seven_days_ago), 'AGENT_BOOKING'].tolist()
        fourteen_days_ago = top_7_dates - timedelta(days=14)
        fourteen_days_values = df.loc[df['Date'].isin(fourteen_days_ago), 'AGENT_BOOKING'].tolist()
        
        twenty_eight_days_ago = top_7_dates - timedelta(days=28)
        twenty_eight_days_values = df.loc[df['Date'].isin(twenty_eight_days_ago), 'AGENT_BOOKING'].tolist()
        fiftysix_days_ago = top_7_dates - timedelta(days=56)
        fiftysix_days_values = df.loc[df['Date'].isin(fiftysix_days_ago), 'AGENT_BOOKING'].tolist()

        same_date_values1 = df.loc[df['Date'].isin(top_7_dates), 'AGENT_CENTER'].tolist()
        seven_days_ago1 = top_7_dates - timedelta(days=7)
        seven_days_values1 = df.loc[df['Date'].isin(seven_days_ago1), 'AGENT_CENTER'].tolist()
        fourteen_days_ago1 = top_7_dates - timedelta(days=14)
        fourteen_days_values1 = df.loc[df['Date'].isin(fourteen_days_ago1), 'AGENT_CENTER'].tolist()
        twenty_eight_days_ago1 = top_7_dates - timedelta(days=28)
        twenty_eight_days_values1 = df.loc[df['Date'].isin(twenty_eight_days_ago1), 'AGENT_CENTER'].tolist()
        fiftysix_days_ago1 = top_7_dates - timedelta(days=56)
        fiftysix_days_values1 = df.loc[df['Date'].isin(fiftysix_days_ago1), 'AGENT_CENTER'].tolist()
        ratio_same=[int(x / y + 0.5)for x, y in zip(same_date_values, same_date_values1)]
        ratio_seven=[int(x / y + 0.5) for x, y in zip(seven_days_values, seven_days_values1)]
        ratio_fourteen=[int(x / y + 0.5) for x, y in zip(fourteen_days_values, fourteen_days_values1)]
        ratio_twenty_eight=[int(x / y + 0.5) for x, y in zip(twenty_eight_days_values, twenty_eight_days_values1)]
        ratio_fifty_six=[int(x / y + 0.5) for x, y in zip(fiftysix_days_values, fiftysix_days_values1)]
        global table_data
        table_data = pd.DataFrame({
            'Date': top_7_dates,
            'Days': day_names,
            'Same Date': same_date_values,
            "-14 Days":fourteen_days_values,
            '-7 Days': seven_days_values,
            "-28 Days":twenty_eight_days_values,
            "-56 Days ":fiftysix_days_values,
            'Same date':same_date_values1,
            '-7 days': seven_days_values1,
            "-14 days":fourteen_days_values1,
            "-28 days":twenty_eight_days_values1,
            "-56 days ":fiftysix_days_values1,
            "same date":ratio_same,
            "-7 day":ratio_seven,
            "14 day":ratio_fourteen,
            "-28 day":ratio_twenty_eight,
            "-56 day":ratio_fifty_six
        })
        total_row = table_data.sum(numeric_only=True)
        total_row['Date'] = 'Total'
        total_row['Days'] = ''


        # Append the total row to the DataFrame
        table_data = table_data._append(total_row, ignore_index=True)
        header = pd.MultiIndex.from_tuples([('AGENT ', 'Date'),
                                            ('', 'Days'),
                                            ("PICKUP COUNT ",'Same Date'),
                                            ( '','-7 Days'),( '','-14 Days'),
                                            ("",'-28 Days'),
                                            ("",'-56 Days'),
                                            ( 'PARTNER COUNT ','Same date'),
                                            ( "",'-7 days'), ( '','-14 Days'),

                                            ( "",'-28 days'),("","-56 days"),
                                            ("RATIO ","same date"),
                                            ( "",'-7 day'), ( '','-14 Days'),

                                            ( "",'-28 day'),("","-56 day")
                                            ])
       



        # Set the multi-index header to the existing table
        table_data.columns = header
        with pd.ExcelWriter(fr'C:\Users\Tanishqa\Desktop\data\{today_date}.xlsx', engine='openpyxl') as writer:
            # Save table1 to Sheet1
            table_data.to_excel(writer, sheet_name=sheet)
        wb=openpyxl.load_workbook(fr"C:\Users\Tanishqa\Desktop\data\{today_date}.xlsx")
        ws=wb["AGENT"]
        rows = ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column)
        for row in rows:
            for cell in row:
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                    top=Side(style='thin'), bottom=Side(style='thin'))

        fill_pattern=PatternFill(patternType="solid", fgColor="6495ED")
        for col_idx in range(2, len(header) + 2):
            column_letter = get_column_letter(col_idx)
            ws[column_letter+"2"].fill=fill_pattern
        ws["B1"].fill=PatternFill(patternType="solid",fgColor="FFF000")
        for i in range(4,15,5):
            column_letter1 = get_column_letter(i)
            ws[column_letter1+"1"].fill=PatternFill(patternType="solid", fgColor="6495ED")
        for idx in range(4,11):
            for j in range(4, len(header) + 1):
                column=get_column_letter(j)+str(idx)
                next_column = get_column_letter(j + 1) + str(idx)

                if ws[column].value>ws[next_column].value:
                    ws[column].fill=PatternFill(patternType="solid",fgColor="90EE90")
                elif ws[column].value==ws[next_column].value:
                    ws[column].fill=PatternFill(patternType="solid",fgColor="FFFFFF")
                else:
                    ws[column].fill=PatternFill(patternType="solid",fgColor="FF0000")
    
       
        wb.save(fr"C:\Users\Tanishqa\Desktop\data\{today_date}.xlsx")
    
        return table_data



def api(df,sheet):
        same_date_values2 = df.loc[df['Date'].isin(top_7_dates), 'API_BOOKING'].tolist()
        seven_days_ago2 = top_7_dates - timedelta(days=7)
        seven_days_values2 = df.loc[df['Date'].isin(seven_days_ago2), 'API_BOOKING'].tolist()
        fourteen_days_ago2 = top_7_dates - timedelta(days=14)
        fourteen_days_values2 = df.loc[df['Date'].isin(fourteen_days_ago2), 'API_BOOKING'].tolist()
        twenty_eight_days_ago2 = top_7_dates - timedelta(days=28)
        twenty_eight_days_values2 = df.loc[df['Date'].isin(twenty_eight_days_ago2), 'API_BOOKING'].tolist()
        fiftysix_days_ago2 = top_7_dates - timedelta(days=56)
        fiftysix_days_values2 = df.loc[df['Date'].isin(fiftysix_days_ago2), 'API_BOOKING'].tolist()

        same_date_values12 = df.loc[df['Date'].isin(top_7_dates), 'API_CENTER'].tolist()
        seven_days_ago12 = top_7_dates - timedelta(days=7)
        seven_days_values12 = df.loc[df['Date'].isin(seven_days_ago12), 'API_CENTER'].tolist()
        fourteen_days_ago12 = top_7_dates - timedelta(days=14)
        fourteen_days_values12 = df.loc[df['Date'].isin(fourteen_days_ago12), 'API_CENTER'].tolist()
        twenty_eight_days_ago12 = top_7_dates - timedelta(days=28)
        twenty_eight_days_values12 = df.loc[df['Date'].isin(twenty_eight_days_ago12), 'API_CENTER'].tolist()
        fiftysix_days_ago12 = top_7_dates - timedelta(days=56)
        fiftysix_days_values12 = df.loc[df['Date'].isin(fiftysix_days_ago12), 'API_CENTER'].tolist()
        ratio_same2=[int(x / y + 0.5)for x, y in zip(same_date_values2, same_date_values12)]
        ratio_seven2=[int(x / y + 0.5) for x, y in zip(seven_days_values2, seven_days_values12)]
        ratio_fourteen2=[int(x / y + 0.5) for x, y in zip(fourteen_days_values2, fourteen_days_values12)]
        ratio_twenty_eight2=[int(x / y + 0.5) for x, y in zip(twenty_eight_days_values2, twenty_eight_days_values12)]
        ratio_fifty_six2=[int(x / y + 0.5) for x, y in zip(fiftysix_days_values2, fiftysix_days_values12)]

        global table_data1
        table_data1 = pd.DataFrame({
            'Date': top_7_dates,
            'Days': day_names,
            'Same Date': same_date_values2,
            '-7 Days': seven_days_values2,
            '-14 Days': fourteen_days_values2,
            "-28 Days":twenty_eight_days_values2,
            "-56 Days ":fiftysix_days_values2,
            'Same date': same_date_values12,
            '-7 days': seven_days_values12,
            '-14 days':fourteen_days_values12,
            "-28 days":twenty_eight_days_values12,
            "-56 days ":fiftysix_days_values12,
            "same date":ratio_same2,
            "-7 day":ratio_seven2,
            "-14 day":ratio_fourteen2,
            "-28 day":ratio_twenty_eight2,
            "-56 day":ratio_fifty_six2
        })
        total_row1 = table_data1.sum(numeric_only=True)
        total_row1['Date'] = 'Total'
        total_row1['Days'] = ''

    # Append the total row to the DataFrame
        table_data1 = table_data1._append(total_row1, ignore_index=True)
  
        
        header = pd.MultiIndex.from_tuples([('API ', 'Date'),
                                            ('', 'Days'),
                                            ("PICKUP COUNT ",'Same Date'),
                                            ( '','-7 Days'),( '','-14 Days'),
                                            ("",'-28 Days'),
                                            ("",'-56 Days'),
                                            ( 'PARTNER COUNT ','Same date'),
                                            ( "",'-7 days'),( "",'-14 days'),
                                            ( "",'-28 days'),("","-56 days"),
                                            ("RATIO ","same date"),
                                            ( "",'-7 day'),( "",'-14 day'),
                                            ( "",'-28 day'),("","-56 day")
                                            ])


        # Set the multi-index header to the existing table
        table_data1.columns = header
        with pd.ExcelWriter(fr'C:\Users\Tanishqa\Desktop\data\{today_date}.xlsx',mode="a", engine='openpyxl') as writer:
               table_data1.to_excel(writer, sheet_name=sheet)
        wb=openpyxl.load_workbook(fr"C:\Users\Tanishqa\Desktop\data\{today_date}.xlsx")
        ws=wb["API"]
        rows = ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column)
        for row in rows:
            for cell in row:
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                    top=Side(style='thin'), bottom=Side(style='thin'))

        fill_pattern=PatternFill(patternType="solid", fgColor="6495ED")
        for col_idx in range(2, len(header) + 2):
            column_letter = get_column_letter(col_idx)
            ws[column_letter+"2"].fill=fill_pattern
        ws["B1"].fill=PatternFill(patternType="solid",fgColor="FFF000")
        for i in range(4,15,5):
            column_letter1 = get_column_letter(i)
            ws[column_letter1+"1"].fill=PatternFill(patternType="solid", fgColor="6495ED")
        for idx in range(4,11):
            for j in range(4, len(header) + 1):
                column=get_column_letter(j)+str(idx)
                next_column = get_column_letter(j + 1) + str(idx)

                if ws[column].value>ws[next_column].value:
                    ws[column].fill=PatternFill(patternType="solid",fgColor="90EE90")
                elif ws[column].value==ws[next_column].value:
                    ws[column].fill=PatternFill(patternType="solid",fgColor="FFFFFF")
                else:
                    ws[column].fill=PatternFill(patternType="solid",fgColor="FF0000")
       
        wb.save(fr"C:\Users\Tanishqa\Desktop\data\{today_date}.xlsx")
    
           

        return table_data1        
        
def frame(df,sheet):
        same_date_values3 = df.loc[df['Date'].isin(top_7_dates), 'IFRAME_BOOKING'].tolist()
        seven_days_ago3 = top_7_dates - timedelta(days=7)
        seven_days_values3= df.loc[df['Date'].isin(seven_days_ago3), 'IFRAME_BOOKING'].tolist()
        fourteen_days_ago3 = top_7_dates - timedelta(days=14)
        fourteen_days_values3= df.loc[df['Date'].isin(fourteen_days_ago3), 'IFRAME_BOOKING'].tolist()
        twenty_eight_days_ago3 = top_7_dates - timedelta(days=28)
        twenty_eight_days_values3 = df.loc[df['Date'].isin(twenty_eight_days_ago3), 'IFRAME_BOOKING'].tolist()
        fiftysix_days_ago3 = top_7_dates - timedelta(days=56)
        fiftysix_days_values3 = df.loc[df['Date'].isin(fiftysix_days_ago3), 'IFRAME_BOOKING'].tolist()

        same_date_values13 = df.loc[df['Date'].isin(top_7_dates), 'IFRAME_CENTER'].tolist()
        seven_days_ago13 = top_7_dates - timedelta(days=7)
        seven_days_values13 = df.loc[df['Date'].isin(seven_days_ago13), 'IFRAME_CENTER'].tolist()
        fourteen_days_ago13 = top_7_dates - timedelta(days=14)
        fourteen_days_values13 = df.loc[df['Date'].isin(fourteen_days_ago13), 'IFRAME_CENTER'].tolist()
        twenty_eight_days_ago13= top_7_dates - timedelta(days=28)
        twenty_eight_days_values13 = df.loc[df['Date'].isin(twenty_eight_days_ago13), 'IFRAME_CENTER'].tolist()
        fiftysix_days_ago13 = top_7_dates - timedelta(days=56)
        fiftysix_days_values13 = df.loc[df['Date'].isin(fiftysix_days_ago13), 'IFRAME_CENTER'].tolist()
        ratio_same3=[int(x / y + 0.5)for x, y in zip(same_date_values3, same_date_values13)]
        ratio_seven3=[int(x / y + 0.5) for x, y in zip(seven_days_values3, seven_days_values13)]
        ratio_fourteen3=[int(x / y + 0.5) for x, y in zip(fourteen_days_values3, fourteen_days_values13)]

        ratio_twenty_eight3=[int(x / y + 0.5) for x, y in zip(twenty_eight_days_values3, twenty_eight_days_values13)]
        ratio_fifty_six3=[int(x / y + 0.5) for x, y in zip(fiftysix_days_values3, fiftysix_days_values13)]

        global table_data2
        table_data2 = pd.DataFrame({
            'Date': top_7_dates,
            'Days': day_names,
            'Same Date': same_date_values3,
            '-7 Days': seven_days_values3,
            '-14 Days': fourteen_days_values3,
            "-28 Days":twenty_eight_days_values3,
            "-56 Days ":fiftysix_days_values3,
            'Same date': same_date_values13,
            '-7 days': seven_days_values13,
            '-14 days': fourteen_days_values13,
            "-28 days":twenty_eight_days_values13,
            "-56 days ":fiftysix_days_values13,
            "same date":ratio_same3,
            "-7 day":ratio_seven3,
            "-14 day":ratio_fourteen3,
            "-28 day":ratio_twenty_eight3,
            "-56 day":ratio_fifty_six3
        })

        total_row2 = table_data2.sum(numeric_only=True)
        total_row2['Date'] = 'Total'
        total_row2['Days'] = ''
        

        # Append the total row to the DataFrame
        table_data2 = table_data2._append(total_row2, ignore_index=True)
        header = pd.MultiIndex.from_tuples([('IFRAME ', 'Date'),
                                            ('', 'Days'),
                                            ("PICKUP COUNT ",'Same Date'),
                                            ( '','-7 Days'),( '','-14 Days'),
                                            ("",'-28 Days'),
                                            ("",'-56 Days'),
                                            ( 'PARTNER COUNT ','Same date'),
                                            ( "",'-7 days'),( '','-14 days'),
                                            ( "",'-28 days'),("","-56 days"),
                                            ("RATIO ","same date"),
                                            ( "",'-7 day'),( '','-14 day'),
                                            ( "",'-28 day'),("","-56 day")
                                            ])
        

        

        # Set the multi-index header to the existing table
        table_data2.columns = header
        with pd.ExcelWriter(fr'C:\Users\Tanishqa\Desktop\data\{today_date}.xlsx', mode='a', engine='openpyxl') as writer:
              table_data2.to_excel(writer, sheet_name=sheet)
        wb=openpyxl.load_workbook(fr"C:\Users\Tanishqa\Desktop\data\{today_date}.xlsx")
        ws=wb["FRAME"]
        rows = ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column)
        for row in rows:
            for cell in row:
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                    top=Side(style='thin'), bottom=Side(style='thin'))

        fill_pattern=PatternFill(patternType="solid", fgColor="6495ED")
        for col_idx in range(2, len(header) + 2):
            column_letter = get_column_letter(col_idx)
            ws[column_letter+"2"].fill=fill_pattern
        ws["B1"].fill=PatternFill(patternType="solid",fgColor="FFF000")  
        for i in range(4,15,5):
            column_letter1 = get_column_letter(i)
            ws[column_letter1+"1"].fill=PatternFill(patternType="solid", fgColor="6495ED")
        for idx in range(4,11):
            for j in range(4, len(header) + 1):
                column=get_column_letter(j)+str(idx)
                next_column = get_column_letter(j + 1) + str(idx)

                if ws[column].value>ws[next_column].value:
                    ws[column].fill=PatternFill(patternType="solid",fgColor="90EE90")
                elif ws[column].value==ws[next_column].value:
                    ws[column].fill=PatternFill(patternType="solid",fgColor="FFFFFF")
                else:
                    ws[column].fill=PatternFill(patternType="solid",fgColor="FF0000")
        wb.save(fr"C:\Users\Tanishqa\Desktop\data\{today_date}.xlsx")
      
        return table_data2       
        
# Display the table with the header
print(agent(df,"AGENT"))
print(api(df,"API"))
print(frame(df,"FRAME"))

def percentage1():
    Heading=["Date","Day","Same Date","-7 Days","-14 Days","-28 Days","-56 Days","Same date","-7 days","-14 Days","-28 days","-56 days","Same date","-7 days","-14 Days","-28 days","-56 days"]
    new_table= pd.DataFrame(columns=Heading)
    row1=table_data.iloc[-1].values
    new_table = new_table._append(pd.Series(row1, index=Heading), ignore_index=True)
    row2=table_data1.iloc[-1].values
    new_table = new_table._append(pd.Series(row2, index=Heading), ignore_index=True)
    row3=table_data2.iloc[-1].values
    new_table = new_table._append(pd.Series(row3, index=Heading), ignore_index=True)
    
    totals= new_table.drop(columns=["Date",'Day']).sum()

    # Calculate the percentage for each row (excluding the "Grand total" row)
    percentage_values = new_table.drop(columns=["Date", "Day"])
    percentage_values = percentage_values.div(totals) * 100
    percentage_values = percentage_values.applymap(lambda value: f"{int(value)}%")
    percentage_values.insert(loc=0,column="Name",value=["AGENT","API","IFRAME"])
    with pd.ExcelWriter(fr'C:\Users\Tanishqa\Desktop\data\{today_date}.xlsx', mode='a', engine='openpyxl') as writer:
        percentage_values.to_excel(writer, sheet_name="percentage")




   



percentage1() 
print("done ,please check your file ")